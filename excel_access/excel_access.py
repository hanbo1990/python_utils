from openpyxl import load_workbook


class ExcelAccessor:

    __file_name = ""
    __workbook = None
    __current_sheet = None

    def __init__(self, excel_name):
        """Init the excel accessor with excel name

        Arguments:
            excel_name {String} -- path to the excel file
        """
        self.__file_name = excel_name
        self.__workbook = load_workbook(filename=excel_name)
        assert self.__workbook is not None

    def go_to_sheet(self, sheet_name, new_if_not_exist=True):
        """Go to certain sheet by name

        Arguments:
            sheet_name {String} -- name of the sheet

        Keyword Arguments:
            new_if_not_exist {bool} -- Create a new sheet if sheet does not exist (default: {True})

        Raises:
            ValueError: [Sheet with name does not exist when new_if_not_exist is False]
        """        
        if sheet_name not in self.__workbook.sheetnames:
            if new_if_not_exist is True:
                self.__workbook.create_sheet(title=sheet_name,
                                             index=len(self.__workbook.sheetnames))
            else:
                raise ValueError(sheet_name + "not found in excel")
        
        self.__current_sheet = self.__workbook[sheet_name]
        assert self.__current_sheet is not None

    def read_cell(self, worksheet_name, row, column):
        """Read a certain cell

        Arguments:
            worksheet_name {String} -- name of the sheet
            row {int} -- row number
            column {String} -- column A-Z

        Returns:
            [String] -- value in the excel
        """        
        self.go_to_sheet(worksheet_name, True)
        cell_name = str(column) + str(row)
        return str(self.__current_sheet[cell_name].value)

    def write_cell(self, worksheet_name, row, column, value):
        """Write value to a certain cell

        Arguments:
            worksheet_name {String} -- name of the sheet
            row {int} -- row number
            column {String} -- column A-Z
            value {String} -- Value to write to the cell
        """        
        self.go_to_sheet(worksheet_name, True)
        cell_name = str(column) + str(row)
        self.__current_sheet[cell_name].value = value
        self.__workbook.save(filename=self.__file_name)

    def get_max_row(self, sheet_name):
        """Get the max row number of certain sheet

        Arguments:
            sheet_name {String} -- name of the sheet

        Returns:
            [int] -- number of rows
        """        
        self.go_to_sheet(sheet_name, True)
        return self.__current_sheet.max_row

    def get_max_column(self, sheet_name):
        """Get the max column number of certain sheet

        Arguments:
            sheet_name {String} -- name of the sheet

        Returns:
            [int] -- number of column
        """        
        self.go_to_sheet(sheet_name, True)
        return self.__current_sheet.max_column

    def merge_cell(self, sheet_name, start_row, end_row, start_column, end_column):
        """Merge cell within a range

        Arguments:
            sheet_name {String} -- name of the sheet
            start_row {[int, String]} -- row number to start merge
            end_row {[int, String]} -- row number to stop merge
            start_column {String} -- column number to start merge
            end_column {String} -- column number to stop merge
        """        
        self.go_to_sheet(sheet_name, True)
        start_column = ord(start_column) - ord("A") + 1
        end_column = ord(end_column) - ord("A") + 1
        self.__current_sheet.merge_cells(start_row=start_row,
                                         end_row=end_row,
                                         start_column=start_column,
                                         end_column=end_column)
        self.__workbook.save(self.__file_name)
